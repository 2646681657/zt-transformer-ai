# -*- coding: utf-8 -*-
"""提取线圈型式 / Sheet1 + VBA 检查 + 外部引用"""
import openpyxl
import sys
import zipfile

sys.stdout.reconfigure(encoding="utf-8")

PATH = r"D:\zt-transformer\zt-transformer-ai-main\SB20-M-630-10计算单.xlsm"

wbf = openpyxl.load_workbook(PATH, data_only=False)
wbv = openpyxl.load_workbook(PATH, data_only=True)

for name in ["线圈型式", "Sheet1"]:
    wsf = wbf[name]
    wsv = wbv[name]
    print("=" * 90)
    print(f"Sheet: {name}")
    print("=" * 90)
    for row in wsf.iter_rows():
        for cf in row:
            if cf.value is None:
                continue
            coord = cf.coordinate
            f = str(cf.value)
            v = wbv[name][coord].value
            if f.startswith("="):
                print(f"{coord}\tF\t{f}\t=> {v}")
            else:
                print(f"{coord}\tV\t{f}")

# 外部链接
print("\n### 外部工作簿链接")
try:
    for i, el in enumerate(wbf._external_links):
        print(f"  [{i+1}] {el.file_link.target}")
except Exception as e:
    print("  读取失败:", e)

# VBA 检查
print("\n### VBA 宏检查")
with zipfile.ZipFile(PATH) as z:
    names = z.namelist()
    vba = [n for n in names if "vba" in n.lower()]
    print("  zip 内 vba 相关文件:", vba if vba else "无")
    if vba:
        for n in vba:
            data = z.read(n)
            print(f"  {n}: {len(data)} bytes")
