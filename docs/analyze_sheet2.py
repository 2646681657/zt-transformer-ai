# -*- coding: utf-8 -*-
"""提取计算单主表所有单元格：坐标 | 公式 | 缓存值"""
import openpyxl
import sys

sys.stdout.reconfigure(encoding="utf-8")

PATH = r"D:\zt-transformer\zt-transformer-ai-main\SB20-M-630-10计算单.xlsm"

wbf = openpyxl.load_workbook(PATH, data_only=False)   # 公式
wbv = openpyxl.load_workbook(PATH, data_only=True)    # 缓存值

for name in ["计算单", "成本测算"]:
    wsf = wbf[name]
    wsv = wbv[name]
    print("=" * 90)
    print(f"Sheet: {name}")
    print("=" * 90)
    for row in wsf.iter_rows():
        for cf in row:
            if cf.value is None:
                continue
            coord = cf.coordinate
            f = str(cf.value)
            v = wsv[coord].value
            if f.startswith("="):
                print(f"{coord}\tF\t{f}\t=> {v}")
            else:
                # 文本/数值
                print(f"{coord}\tV\t{f}")

print("\n### 定义的名称 (Defined Names)")
try:
    for n, dn in wbf.defined_names.items():
        print(f"  {n} = {dn.value}")
except Exception as e:
    print("  (无)", e)
