# -*- coding: utf-8 -*-
"""解析计算单 xlsm：输出工作表结构概览"""
import openpyxl
import sys

sys.stdout.reconfigure(encoding="utf-8")

PATH = r"D:\zt-transformer\zt-transformer-ai-main\SB20-M-630-10计算单.xlsm"

# 保留公式读取
wb = openpyxl.load_workbook(PATH, data_only=False, keep_vba=True)

print("=" * 70)
print(f"工作表数量: {len(wb.sheetnames)}")
print("=" * 70)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n### Sheet: {name!r}")
    print(f"    尺寸: {ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
    print(f"    可见性: {ws.sheet_state}")
    # 合并单元格数量
    print(f"    合并单元格数: {len(ws.merged_cells.ranges)}")

# 值读取（data_only 需要 Excel 缓存过）
wb2 = openpyxl.load_workbook(PATH, data_only=True)
print("\n" + "=" * 70)
print("各 sheet 非空单元格统计（值缓存）")
print("=" * 70)
for name in wb2.sheetnames:
    ws = wb2[name]
    cnt = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                cnt += 1
    print(f"  {name!r}: {cnt} 个非空单元格")
